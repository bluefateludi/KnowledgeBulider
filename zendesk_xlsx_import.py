#!/usr/bin/env python3
import argparse, json, os, sys, subprocess
def rq(b,e,t,p,m="GET",d=None):
    u=f"{b}{p}"
    if m!="GET" or d is not None:
        cmd=["curl","-sS","-u",f"{e}/token:{t}","-H","Content-Type: application/json","-X",m]
        if d is not None:
            cmd+=["-d",json.dumps(d,ensure_ascii=False)]
        cmd+=[u]
    else:
        cmd=["curl","-sS","-u",f"{e}/token:{t}",u]
    res=subprocess.run(cmd,capture_output=True,text=True)
    out=res.stdout.strip()
    try:
        return json.loads(out) if out else {}
    except Exception:
        return {"_error":True,"parse_error":out[:200]}
def gc(b,e,t):return rq(b,e,t,"/api/v2/help_center/categories.json","GET",None)
def cc(b,e,t,n,d=""):return rq(b,e,t,"/api/v2/help_center/categories.json","POST",{"category":{"name":n,"description":d}})
def gs(b,e,t,c):return rq(b,e,t,f"/api/v2/help_center/categories/{c}/sections.json","GET",None)
def cs(b,e,t,c,n,d=""):return rq(b,e,t,f"/api/v2/help_center/categories/{c}/sections.json","POST",{"section":{"name":n,"description":d}})
def ca(b,e,t,s,ti,bd,pg,us,lc,dr):return rq(b,e,t,f"/api/v2/help_center/sections/{s}/articles.json","POST",{"article":{"title":ti,"body":bd,"permission_group_id":pg,"user_segment_id":us,"locale":lc,"draft":dr}})
def ecid(b,e,t,n):
    x=gc(b,e,t)
    if "_error" not in x:
        for c in x.get("categories",[]): 
            if c.get("name")==n:return c.get("id")
    y=cc(b,e,t,n)
    if "_error" in y:
        x=gc(b,e,t)
        for c in x.get("categories",[]):
            if c.get("name")==n:return c.get("id")
        return None
    return y.get("category",{}).get("id")
def esid(b,e,t,c,n):
    x=gs(b,e,t,c)
    if "_error" not in x:
        for s in x.get("sections",[]):
            if s.get("name")==n:return s.get("id")
    y=cs(b,e,t,c,n)
    if "_error" in y:
        x=gs(b,e,t,c)
        for s in x.get("sections",[]):
            if s.get("name")==n:return s.get("id")
        return None
    return y.get("section",{}).get("id")
def read_xlsx(path,sheet=None):
    import openpyxl
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
    ws=wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows=list(ws.iter_rows(values_only=True))
    if not rows:return []
    hdr=[str(c).strip() if c is not None else "" for c in rows[0]]
    def idx(names):
        for i,v in enumerate(hdr):
            if v in names:return i
        return -1
    ci=idx(["类别","Category","分类"])
    si=idx(["组别","Section","分组"])
    ti=idx(["文章标题","Title"])
    bi=idx(["文章内容","Body","内容"])
    if min(ci,si,ti,bi)<0: 
        raise RuntimeError("表头需包含: 类别, 组别, 文章标题, 文章内容")
    out=[]
    for row in rows[1:]:
        if row is None: continue
        get=lambda i:(str(row[i]).strip() if i>=0 and i<len(row) and row[i] is not None else "")
        c=get(ci); s=get(si); t=get(ti); b=get(bi)
        if not t: continue
        out.append((c,s,t,b))
    return out
def main():
    p=argparse.ArgumentParser()
    p.add_argument("--subdomain",required=True)
    p.add_argument("--email",required=True)
    p.add_argument("--api-token",required=True)
    p.add_argument("--xlsx",required=True)
    p.add_argument("--sheet")
    p.add_argument("--permission-group-id",type=int,default=15285357366159)
    p.add_argument("--locale",default="zh-cn")
    p.add_argument("--draft",choices=["true","false"],default="false")
    p.add_argument("--user-segment-id",default="null")
    args=p.parse_args()
    base=f"https://{args.subdomain}.zendesk.com"
    dr=True if args.draft.lower()=="true" else False
    us=None if args.user_segment_id=="null" else int(args.user_segment_id)
    rows=read_xlsx(args.xlsx,args.sheet)
    for i,(cat,sec,title,body) in enumerate(rows,1):
        if not cat or not sec:
            print(json.dumps({"row":i,"status":"skip","reason":"缺少分类或组别"},ensure_ascii=False))
            continue
        if "<" not in body: body=f"<p>{body}</p>"
        cid=ecid(base,args.email,args.api_token,cat)
        if not cid:
            print(json.dumps({"row":i,"status":"fail","reason":"分类失败","cat":cat},ensure_ascii=False))
            continue
        sid=esid(base,args.email,args.api_token,cid,sec)
        if not sid:
            print(json.dumps({"row":i,"status":"fail","reason":"组别失败","section":sec},ensure_ascii=False))
            continue
        created=ca(base,args.email,args.api_token,sid,title,body,args.permission_group_id,us,args.locale,dr)
        if "_error" in created:
            print(json.dumps({"row":i,"status":"fail","reason":created},ensure_ascii=False))
            continue
        art=created.get("article",{})
        print(json.dumps({"row":i,"status":"ok","article_id":art.get("id"),"html_url":art.get("html_url"),"title":title},ensure_ascii=False))
if __name__=="__main__": 
    main()
